"""Export a valuation run to a formatted Excel workbook.

Produces one sheet per statement plus a DCF sheet and a sensitivity sheet,
using the industry-standard convention of blue input cells and black
calculated cells.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

from ..data.base import CompanyFinancials
from ..model.dcf import DCFResult

_BLUE = "0000FF"
_HDR_FILL = "1F3864"
_SUB_FILL = "D9E1F2"


def export(path: str | Path, fin: CompanyFinancials,
           income: pd.DataFrame, balance: pd.DataFrame, cashflow: pd.DataFrame,
           fcf: pd.Series, result: DCFResult) -> Path:
    """Write the workbook to ``path`` and return the resolved path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export.") from exc

    title_font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    label_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)
    blue_font = Font(name="Arial", size=10, color=_BLUE)
    hdr_fill = PatternFill("solid", fgColor=_HDR_FILL)
    sub_fill = PatternFill("solid", fgColor=_SUB_FILL)
    money = '#,##0;(#,##0)'
    money2 = '#,##0.00;(#,##0.00)'
    pct = '0.0%;(0.0%)'

    wb = Workbook()

    def write_statement(name: str, frame: pd.DataFrame, fmt: str = money) -> None:
        ws = wb.create_sheet(name)
        ws["A1"] = f"{fin.ticker} - {name} (millions {fin.currency})"
        ws["A1"].font = title_font
        ws["A1"].fill = hdr_fill
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=frame.shape[1] + 1)
        ws["A3"] = "Line item"
        ws["A3"].font = hdr_font
        ws["A3"].fill = hdr_fill
        for j, col in enumerate(frame.columns, start=2):
            cell = ws.cell(row=3, column=j, value=str(col))
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="right")
        for i, (label, values) in enumerate(frame.iterrows(), start=4):
            ws.cell(row=i, column=1, value=str(label)).font = label_font
            for j, value in enumerate(values, start=2):
                cell = ws.cell(row=i, column=j, value=float(value))
                cell.font = body_font
                cell.number_format = fmt
        ws.column_dimensions["A"].width = 26
        for j in range(2, frame.shape[1] + 2):
            ws.column_dimensions[chr(64 + j)].width = 14

    # Cover / DCF summary sheet first.
    cover = wb.active
    cover.title = "DCF Summary"
    cover["A1"] = f"{fin.name} ({fin.ticker}) - DCF Valuation"
    cover["A1"].font = title_font
    cover["A1"].fill = hdr_fill
    cover.merge_cells("A1:C1")
    cover.column_dimensions["A"].width = 30
    cover.column_dimensions["B"].width = 16

    summary_rows = [
        ("Cost of equity", result.cost_of_equity, pct),
        ("After-tax cost of debt", result.cost_of_debt_after_tax, pct),
        ("WACC", result.wacc, pct),
        ("PV of explicit FCF", result.pv_explicit_fcf, money),
        ("PV of terminal value", result.pv_terminal_value, money),
        ("Terminal value % of EV", result.terminal_value_pct, pct),
        ("Enterprise value", result.enterprise_value, money),
        ("Less: net debt", result.net_debt, money),
        ("Equity value", result.equity_value, money),
        ("Shares outstanding (mm)", result.shares_outstanding, money),
        ("Implied share price", result.implied_share_price, money2),
        ("Current share price", result.current_share_price, money2),
        ("Upside / (downside)", result.upside, pct),
    ]
    cover["A3"] = "Metric"
    cover["B3"] = "Value"
    for c in ("A3", "B3"):
        cover[c].font = hdr_font
        cover[c].fill = hdr_fill
    for i, (label, value, fmt) in enumerate(summary_rows, start=4):
        cover.cell(row=i, column=1, value=label).font = label_font
        cell = cover.cell(row=i, column=2, value=float(value))
        cell.number_format = fmt
        cell.font = body_font

    write_statement("Income Statement", income)
    write_statement("Balance Sheet", balance)
    write_statement("Cash Flow", cashflow)
    write_statement("FCFF", fcf.to_frame("Unlevered FCF").T)

    # Sensitivity sheet.
    sens = wb.create_sheet("Sensitivity")
    sens["A1"] = "Implied share price: WACC (rows) x terminal growth (cols)"
    sens["A1"].font = title_font
    sens["A1"].fill = hdr_fill
    sens.merge_cells("A1:G1")
    sens["A3"] = "WACC \\ g"
    sens["A3"].font = hdr_font
    sens["A3"].fill = hdr_fill
    grid = result.sensitivity
    for j, col in enumerate(grid.columns, start=2):
        cell = sens.cell(row=3, column=j, value=str(col))
        cell.font = hdr_font
        cell.fill = hdr_fill
    for i, (idx, values) in enumerate(grid.iterrows(), start=4):
        head = sens.cell(row=i, column=1, value=str(idx))
        head.font = hdr_font
        head.fill = sub_fill
        for j, value in enumerate(values, start=2):
            cell = sens.cell(row=i, column=j,
                             value=float(value) if pd.notna(value) else None)
            cell.number_format = money2
            cell.font = body_font
    sens.column_dimensions["A"].width = 12

    out = Path(path)
    wb.save(out)
    return out.resolve()
